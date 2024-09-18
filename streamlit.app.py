import streamlit as st
import pandas as pd
import pymysql
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

# Function to connect to MySQL and execute a query
def execute_query(query, params=None):
    try:
        connection = pymysql.connect(host='endpoint', user='root', passwd='******2021', database='project')
        df = pd.read_sql(query, connection, params=params)
        return df
    except Exception as e:
        st.error(f"Error: {e}")
        return pd.DataFrame()
    finally:
        connection.close()

# Function to generate SQL query based on filters
def generate_query(bustypes, routes, amenities, price_min, price_max, rating_min, rating_max, availability):
    query = "SELECT * FROM redbus WHERE 1=1"
    
    if bustypes:
        bustype_str = ",".join(["%s"] * len(bustypes))
        query += f" AND Bus_Type IN ({bustype_str})"
    if routes:
        route_str = ",".join(["%s"] * len(routes))
        query += f" AND Bus_Route IN ({route_str})"
    if amenities:
        amenities_str = " AND ".join([f"Amenities LIKE '%{amenity}%'" for amenity in amenities])
        query += f" AND ({amenities_str})"
    if price_min is not None:
        query += f" AND Fare >= {price_min}"
    if price_max is not None:
        query += f" AND Fare <= {price_max}"
    if rating_min is not None:
        query += f" AND Rating >= {rating_min}"
    if rating_max is not None:
        query += f" AND Rating <= {rating_max}"
    if availability is not None:
        query += f" AND Seat_Availability >= {availability}"
    
    return query

# Define the List2xlsx class for exporting and formatting
class List2xlsx:
    
    @staticmethod
    def to_excel(df, filename):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
        
        output.seek(0)
        with open(filename, 'wb') as f:
            f.write(output.read())
        
        # Add formatting
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = font
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
        wb.save(filename)

# Streamlit application
def main():
    # Custom CSS for styling
    st.markdown("""
        <style>
            .main {
                background-color: #ffcccc; /* Light red background */
                color: #333333; /* Dark text color */
            }
            .sidebar {
                background-color: #ffffff; /* White sidebar for contrast */
                color: #333333; /* Dark text color */
            }
            .stButton > button {
                background-color: #cc0000; /* Darker red for buttons */
                color: #ffffff; /* White text */
            }
            .stTextInput > div > input, .stNumberInput > div > div > input, .stSlider > div > div > input {
                background-color: #ffffff; /* White input fields */
                color: #333333; /* Dark text color */
            }
            .stDownloadButton > button {
                background-color: #cc0000; /* Darker red for download button */
                color: #ffffff; /* White text */
            }
            .stMarkdown {
                color: #333333; /* Dark text color */
            }
            .css-1d391kg { color: #cc0000; } /* Title color */
            .card { border: 1px solid #e3e3e3; border-radius: 10px; padding: 10px; margin: 10px; }
            .card-header { background-color: #cc0000; color: #ffffff; padding: 10px; border-radius: 10px 10px 0 0; }
            .card-body { padding: 10px; }
            .header-banner {
                background-color: #cc0000; /* Banner color */
                color: #ffffff; /* Text color */
                padding: 10px;
                text-align: center;
                font-size: 24px;
                font-weight: bold;
            }
            .top-options {
                margin: 10px;
                text-align: center;
            }
            .top-options a {
                text-decoration: none;
                color: #cc0000;
                font-size: 18px;
                margin: 0 15px;
            }
            .top-options a:hover {
                text-decoration: underline;
            }
        </style>
        """, unsafe_allow_html=True)

    # Add the highlighter/banner at the top
    st.markdown("""
        <div class="header-banner">
            India's No. 1 Online Bus Ticket Booking Site
        </div>
        """, unsafe_allow_html=True)
    
    # Add top options
    st.markdown("""
        <div class="top-options">
            <a href="#download">Download</a>
            <a href="#view-tickets">View Tickets</a>
        </div>
        """, unsafe_allow_html=True)
    
    st.title('RedBus')

    # Search Bus functionality
    st.subheader('Search Bus')
    from_location = st.text_input('From')
    to_location = st.text_input('To')

    # Sidebar filters
    st.sidebar.header('Filter Options')

    bustypes_df = execute_query("SELECT DISTINCT Bus_Type FROM redbus")
    routes_df = execute_query("SELECT DISTINCT Bus_Route FROM redbus")
    
    bustypes = bustypes_df['Bus_Type'].tolist() if 'Bus_Type' in bustypes_df.columns else []
    routes = routes_df['Bus_Route'].tolist() if 'Bus_Route' in routes_df.columns else []

    selected_bustypes = st.sidebar.multiselect("Bus Type", bustypes)
    selected_routes = st.sidebar.multiselect("Bus Route", routes)
    price_range = st.sidebar.slider("Price Range", 0, 5000, (0, 5000))
    rating_range = st.sidebar.slider("Star Rating", 0.0, 5.0, (0.0, 5.0))
    availability = st.sidebar.number_input("Minimum Availability", min_value=0, value=0, step=1)

    amenities_options = [
        "WIFI", "Water Bottle", "Blankets", "Charging Point",
        "Track My Bus", "Emergency Contact Number", "Toilet", "Bed Sheet"
    ]
    selected_amenities = [amenity for amenity in amenities_options if st.sidebar.checkbox(amenity, False)]

    # Generate and execute query
    query = generate_query(
        selected_bustypes,
        selected_routes,
        selected_amenities,
        price_range[0],
        price_range[1],
        rating_range[0],
        rating_range[1],
        availability
    )
    params = selected_bustypes + selected_routes
    filtered_df = execute_query(query, params)

    # Display filtered data
    st.subheader('Filtered Data')
    if not filtered_df.empty:
        st.dataframe(filtered_df)

        # Export to Excel
        st.write("### Export to Excel")
        if st.button("Download Excel File"):
            filename = 'filtered_data.xlsx'
            List2xlsx.to_excel(filtered_df, filename)
            with open(filename, 'rb') as f:
                st.download_button(
                    label="Download Excel",
                    data=f,
                    file_name=filename,
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
    else:
        st.write("No data found with the selected filters.")

    # Help and Account Sections
    st.markdown("---")
    with st.expander("Help"):
        st.write("For assistance with using this app, please contact support@example.com.")
    
    with st.expander("Account"):
        st.write("Manage your account settings here.")

    st.markdown("---")
    st.markdown("Created with Streamlit and MySQL")
    st.markdown("For more information, visit [Streamlit Documentation](https://docs.streamlit.io)")

if __name__ == '__main__':
    main()
